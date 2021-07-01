# -*- coding: utf-8 -*-
"""
Created on Tue Jun 22 17:13:19 2021

@author: calle
"""
import openpyxl

def export_solution_rfep(excel_input_file,
                         excel_output_file,
                        scenario_name,
                        output_solve,
                        b_domain_reduction = False,
                        b_print_solution_detail = False,
                        b_print_location = False,
                        b_print_statistics = False,
                        total_time=0,
                        sVehiclesPaths = [],
                        sOriginalStationsPotential=[],
                        sSequenceNodesNodesVehiclesPaths=[],
                        sDestinationVehiclePath=[],
                        sStationsPaths=set(),
                        sOriginalStationsOwn=[],
                        sStationsVehiclesPaths=[],
                        sSuppliersRanges=[],
                        pStartInventory=0,
                        pConsumptionRate=0,
                        pDistance=0,
                        pSubDistance=0,
                        pConsumptionMainRoute=0,
                        pDistanceOOP=0,
                        pConsumptionOOP=0,
                        pQuantityVehicles=0,
                        pVariableCost=0,
                        pOpportunityCost=0,
                        pLocationCost=0,
                        pStationCapacity=0,
                        pStationUnitCapacity=0,
                        pCostUnitCapacity=0,
                        pPrice=0,
                        pDiscount=0,
                        ):
      #Assign the stats of the solution to variables with name (code readability)
      status = output_solve[0]
      ovInventory = output_solve[1]
      ovRefuelQuantity = output_solve[2]
      ovRefuel = output_solve[3]
      ovQuantityUnitsCapacity = output_solve[4]
      ovLocate = output_solve[5]
      ovQuantityPurchased = output_solve[6]
      ovQuantityPurchasedRange = output_solve[7]
      ovPurchasedRange = output_solve[8]
      oTotalRefuellingCost = output_solve[9]
      oTotalLocationCost = output_solve[10]
      oTotalDiscount = output_solve[11]
      oTotalCost = output_solve[12]
      n_constraints = output_solve[13]
      n_variables = output_solve[14]
      n_integer_variables = output_solve[15]
      n_binary_variables = output_solve[16]
      model_fingerprint = output_solve[17]
      model_runtime = output_solve[18]
      model_MIPGap = output_solve[19]
      
      #excel_file must be the complete path
      output_file = excel_output_file
      #Open Excel workbook
      workbook = openpyxl.load_workbook(output_file)
      scenario = scenario_name  
      sheet_name = 'oNodesNodesVehiclesPaths'
      ws = workbook[sheet_name]
      index_row = ws.max_row    
  

       
      if b_print_solution_detail:
          for (i,j,v,p) in sSequenceNodesNodesVehiclesPaths:
             if (j,p) in sStationsPaths:
                    index_row=index_row +1
                    ws.cell(row=index_row, column = 1, value=scenario)
                    ws.cell(row=index_row, column = 2, value=i)
                    ws.cell(row=index_row, column = 3, value=j)
                    ws.cell(row=index_row, column = 4, value=v)
                    ws.cell(row=index_row, column = 5, value=p)
                    ws.cell(row=index_row, column = 6, value=ovInventory[j,v,p])
                    ws.cell(row=index_row, column = 7, value=ovRefuelQuantity[j,v,p])
                    ws.cell(row=index_row, column = 8, value=ovRefuel[j,v,p])
                    ws.cell(row=index_row, column = 9, value=pStartInventory[v,p])
                    ws.cell(row=index_row, column = 10, value=pConsumptionRate[v])
                    #if a domain reduction procedure was applied, the distance 
                    #changed of domain
                    if b_domain_reduction:
                        ws.cell(row=index_row, column = 11, value=pSubDistance[i,j,v,p])
                    else:
                        ws.cell(row=index_row, column = 11, value=pDistance[i,j,p])   
                    
                    #trhe distance changed of index after the domain reduction. So the original distance wont fit
                    #
                    
                    ws.cell(row=index_row, column = 12, value=pConsumptionMainRoute[i,j,v,p])
                    ws.cell(row=index_row, column = 13, value=pDistanceOOP[j,p])
                    ws.cell(row=index_row, column = 14, value=pConsumptionOOP[j,v,p])
                    ws.cell(row=index_row, column = 15, value=pPrice[j])
                    ws.cell(row=index_row, column = 16, value=0) #this is a placeholder in case I need this to run single corridor
                    ws.cell(row=index_row, column = 17, value=pQuantityVehicles[v,p])
                    ws.cell(row=index_row, column = 18, value=pVariableCost[v])
                    ws.cell(row=index_row, column = 19, value=pOpportunityCost[v])
                    ws.cell(row=index_row, column = 20, value=excel_input_file)
                      
      #Print location variables
      if b_print_location:
          
          sheet_name = 'oOriginalStationsOwn'
          ws = workbook[sheet_name]
          index_row = ws.max_row    
            
          for i in sOriginalStationsOwn:
             index_row = index_row+1
             ws.cell(row=index_row, column = 1, value = scenario)
             ws.cell(row=index_row, column = 2, value = i)
             if i in sOriginalStationsPotential:
                    ws.cell(row=index_row, column = 3, value = ovLocate[i])
                    ws.cell(row=index_row, column = 7, value = pLocationCost[i])
             else:
                    ws.cell(row=index_row, column = 3, value = 0)
                    ws.cell(row=index_row, column = 7, value = 0)
             ws.cell(row=index_row, column = 4, value = ovQuantityUnitsCapacity[i])
             ws.cell(row=index_row, column = 5, value = pStationCapacity[i])
             ws.cell(row=index_row, column = 6, value = pStationUnitCapacity[i])
             ws.cell(row=index_row, column = 8, value = pCostUnitCapacity[i])
             ws.cell(row=index_row, column = 9, value = excel_input_file)
        #Print total cost
      sheet_name = 'oTotalCost'
      ws = workbook[sheet_name]
      index_row = ws.max_row + 1
      ws.cell(row = index_row, column = 1, value = scenario)
      ws.cell(row = index_row, column = 2, value = oTotalRefuellingCost)
      ws.cell(row = index_row, column = 3, value = oTotalLocationCost)
      ws.cell(row = index_row, column = 4, value = oTotalDiscount)
      ws.cell(row = index_row, column = 5, value = oTotalCost)
      ws.cell(row = index_row, column = 6, value = excel_input_file)
      
      
      if b_print_statistics:
          #calculate statistics
          n_vehicles = len({v for (v,p) in sVehiclesPaths})
          n_paths = len({p for (v,p) in sVehiclesPaths})
          n_avg_stations_path = len(sStationsPaths)/n_paths
          n_candidate_locations = len(sOriginalStationsPotential)
          
          #print solution statistics
          sheet_name = 'oScenarioStats'
          ws = workbook[sheet_name]
          index_row = ws.max_row + 1
          ws.cell(row = index_row, column = 1, value = excel_input_file)
          ws.cell(row = index_row, column = 2, value = scenario_name)
          ws.cell(row = index_row, column = 3, value = n_vehicles)
          ws.cell(row = index_row, column = 4, value = n_paths)
          ws.cell(row = index_row, column = 5, value = n_avg_stations_path)
          ws.cell(row = index_row, column = 6, value = n_candidate_locations)
          
          ws.cell(row = index_row, column = 12, value = model_runtime)
          ws.cell(row = index_row, column = 13, value = n_constraints)
          ws.cell(row = index_row, column = 14, value = n_variables)
          ws.cell(row = index_row, column = 15, value = n_integer_variables)
          ws.cell(row = index_row, column = 16, value = n_binary_variables)
          ws.cell(row = index_row, column = 17, value = model_fingerprint)
          ws.cell(row = index_row, column = 18, value = model_MIPGap)
          ws.cell(row = index_row, column = 19, value = total_time)
        
      workbook.save(output_file)    
    

